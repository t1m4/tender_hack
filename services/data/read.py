import json
import time

from openpyxl import load_workbook

from core.models import CTE, CTE_product, Contract, ContractCTE


def save_cte(filename):
    start_time = time.time()
    wb2 = load_workbook(filename)
    ws = wb2.active

    rows = ws.rows
    next(rows)
    k = 1
    cte_objects = []
    for row in rows:
        if k % 10000 == 0:
            print(k)
        row_result = [i.value for i in row]
        cte_instance = CTE(cte_id=row_result[0], name=row_result[1], category=row_result[2],
                           code_kphz=row_result[3])
        cte_objects.append(cte_instance)
        k += 1
    CTE.objects.bulk_create(cte_objects, batch_size=10000)
    print('2', time.time() - start_time)


def save_cte_product(filename):
    start_time = time.time()
    wb2 = load_workbook(filename)
    ws = wb2.active

    rows = ws.rows
    next(rows)
    k = 1
    cte_product_objects = []
    for row in rows:
        if k % 10000 == 0:
            print(k)
        row_result = [i.value for i in row]
        cte_instance = CTE.objects.filter(cte_id=row_result[0], name=row_result[1], category=row_result[2],
                                          code_kphz=row_result[3]).first()
        # add products if they are exists
        if row_result[-1]:
            row_result[-1] = json.loads(row_result[-1])
            for product in row_result[-1]:
                cte_product_instance = CTE_product(name=product.get('name'), cte_product_id=product.get('Id'),
                                                   value=product.get('Value'),
                                                   cte=cte_instance)
                cte_product_objects.append(cte_product_instance)
            CTE_product.objects.bulk_create(cte_product_objects)
        k += 1
    print(len(cte_product_objects))
    # CTE_product.objects.bulk_create(cte_product_objects, batch_size=10000)

    print('2', time.time() - start_time)


def save_contract(filename):
    start_time = time.time()
    wb2 = load_workbook(filename)
    ws = wb2.active

    rows = ws.rows
    next(rows)
    k = 1
    contract_objects = []
    cte_product_objects = []
    contract_cte_objects = []

    for row in rows:
        if k % 10000 == 0:
            print(k)
        try:
            row_result = [i.value for i in row]
            row_result[-1] = json.loads(row_result[-1])

            contract_instance = Contract.objects.create(number=row_result[0], public_date=row_result[1], close_date=row_result[2],
                                    price=row_result[3],
                                    сustomer_inn=row_result[4],
                                    сustomer_kpp=row_result[5],
                                    customer_name=row_result[6],
                                    supplier_inn=row_result[7],
                                    supplier_kpp=row_result[8],
                                    supplier_name=row_result[9],
                                    )
            contract_objects.append(contract_instance)

            for i in row_result[-1]:
                id = i.get("Id")
                quantity = i.get("Quantity")
                amount = i.get("Amount")
                if id:
                    try:
                        contract_cte_instance = ContractCTE.objects.create(contract=contract_instance, quantity=quantity, amount=amount)
                        contract_cte_objects.append(contract_cte_instance)
                        cte_instance = CTE.objects.get(cte_id=id)
                        cte_instance.contract_cte = contract_cte_instance
                        cte_instance.save()

                    except CTE.DoesNotExist:
                        print('break')
                        break
        except Exception as e:
            print('{} {} '.format(k, e))
        k += 1
    print(len(contract_objects))
    # Contract.objects.bulk_create(contract_objects)

    print('2', time.time() - start_time)


def save_contract_cte(filename):
    start_time = time.time()
    wb2 = load_workbook(filename)
    ws = wb2.active

    rows = ws.rows
    next(rows)
    k = 1
    contract_objects = []
    cte_product_objects = []

    for row in rows:
        if k % 10000 == 0:
            print(k)
        try:
            row_result = [i.value for i in row]
            row_result[-1] = json.loads(row_result[-1])

            contract_instance = Contract.objects.filter(number=row_result[0], public_date=row_result[1], close_date=row_result[2],
                                    price=row_result[3],
                                    сustomer_inn=row_result[4],
                                    сustomer_kpp=row_result[5],
                                    customer_name=row_result[6],
                                    supplier_inn=row_result[7],
                                    supplier_kpp=row_result[8],
                                    supplier_name=row_result[9],
                                    ).first()
            contract_objects.append(contract_instance)

            # for i in row_result[-1]:
            #     id = i.get("Id")
            #     quantity = i.get("Quantity")
            #     amount = i.get("Amount")
            #     if id:
            #         try:
            #             cte_instance = CTE.objects.get(cte_id=id)
            #             print(cte_instance.id, contract_instance.id)
            #             # cte_instance.contract = contract_instance
            #             # cte_instance.save()
            #         except CTE.DoesNotExist:
            #             print('break')
            #             break
        except Exception as e:
            print('{} {} '.format(k, e))
        k += 1
    print(len(contract_objects))
    Contract.objects.bulk_create(contract_objects)

    print('2', time.time() - start_time)


if __name__ == '__main__':
    # print(read("Contract.xlsx"))
    # print(read_xsxl("Contract.xlsx"))
    # print(read_cte("CTE.xlsx"))
    print(save_cte("cte_test.xlsx"))
    # print(read_xsxl("cte_hard_test.xlsx"))
